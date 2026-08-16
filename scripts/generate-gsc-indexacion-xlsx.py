"""Genera Excel con listas GSC — estructura exacta solicitada por el usuario."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "exports" / "GSC-solicitud-indexacion-livendia-2026-08-16.xlsx"

PILLAR_VENDER_BASE = "https://livendia.com/vender-piso-sin-inmobiliaria"

Row = tuple[str, str]  # url, categoria

SHEETS: list[tuple[str, list[Row]]] = [
    (
        "Acomp alquiler local",
        [
            ("https://livendia.com/servicios/acompanamiento-alquiler-local", "Acompañamiento alquiler local"),
            (
                "https://livendia.com/servicios/acompanamiento-alquiler-local/hospitalet-de-llobregat",
                "Acompañamiento alquiler local",
            ),
            ("https://livendia.com/servicios/acompanamiento-alquiler-local/madrid", "Acompañamiento alquiler local"),
        ],
    ),
    (
        "Alquiler local metro BCN",
        [
            ("https://livendia.com/servicios/contrato-alquiler-local/barcelona-gracia", "Contrato alquiler local (metro BCN)"),
            (
                "https://livendia.com/servicios/contrato-alquiler-local/barcelona-les-corts",
                "Contrato alquiler local (metro BCN)",
            ),
            (
                "https://livendia.com/servicios/contrato-alquiler-local/cornella-de-llobregat",
                "Contrato alquiler local (metro BCN)",
            ),
            (
                "https://livendia.com/servicios/contrato-alquiler-local/hospitalet-de-llobregat",
                "Contrato alquiler local (metro BCN)",
            ),
        ],
    ),
    (
        "Temporada local",
        [
            ("https://livendia.com/servicios/contrato-alquiler-temporada-local/malaga", "Temporada local"),
            ("https://livendia.com/servicios/contrato-alquiler-temporada-local/valencia", "Temporada local"),
        ],
    ),
    (
        "Arras local",
        [
            ("https://livendia.com/servicios/contrato-arras-local/malaga", "Arras local"),
            ("https://livendia.com/servicios/contrato-arras-local/palma", "Arras local"),
            ("https://livendia.com/servicios/contrato-arras-local/sant-cugat-del-valles", "Arras local"),
            ("https://livendia.com/servicios/contrato-arras-local/sevilla", "Arras local"),
            ("https://livendia.com/servicios/contrato-arras-local/zaragoza", "Arras local"),
        ],
    ),
    (
        "Entre particulares",
        [
            ("https://livendia.com/servicios/contrato-entre-particulares-local", "Contrato entre particulares (clúster)"),
            (
                "https://livendia.com/servicios/contrato-entre-particulares-local/alicante",
                "Contrato entre particulares (clúster)",
            ),
            (
                "https://livendia.com/servicios/contrato-entre-particulares-local/bilbao",
                "Contrato entre particulares (clúster)",
            ),
            (
                "https://livendia.com/servicios/contrato-entre-particulares-local/cordoba",
                "Contrato entre particulares (clúster)",
            ),
            (
                "https://livendia.com/servicios/contrato-entre-particulares-local/gijon",
                "Contrato entre particulares (clúster)",
            ),
            (
                "https://livendia.com/servicios/contrato-entre-particulares-local/granada",
                "Contrato entre particulares (clúster)",
            ),
            (
                "https://livendia.com/servicios/contrato-entre-particulares-local/valladolid",
                "Contrato entre particulares (clúster)",
            ),
            (
                "https://livendia.com/servicios/contrato-entre-particulares-local/zaragoza",
                "Contrato entre particulares (clúster)",
            ),
        ],
    ),
    (
        "Revision post-arras",
        [
            ("https://livendia.com/servicios/revision-documental-post-arras", "Revisión post-arras"),
            ("https://livendia.com/servicios/revision-documental-post-arras/malaga", "Revisión post-arras"),
            ("https://livendia.com/servicios/revision-documental-post-arras/valencia", "Revisión post-arras"),
            (
                "https://livendia.com/servicios/revision-documental-post-arras/hospitalet-de-llobregat",
                "Revisión post-arras",
            ),
            (
                "https://livendia.com/servicios/revision-documental-post-arras/cornella-de-llobregat",
                "Revisión post-arras",
            ),
            ("https://livendia.com/servicios/revision-documental-post-arras/les-corts", "Revisión post-arras"),
        ],
    ),
    (
        "Venta y pilares",
        [
            ("https://livendia.com/servicios/reserva-de-compra", "Venta / vender sin inmobiliaria"),
            ("https://livendia.com/servicios/revision-contrato-alquiler", "Venta / vender sin inmobiliaria"),
            ("https://livendia.com/servicios/vender-piso-sin-agencia", "Venta / vender sin inmobiliaria"),
            ("https://livendia.com/servicios/vender-piso-sin-agencia-granada", "Venta / vender sin inmobiliaria"),
            ("https://livendia.com/servicios/vender-piso-sin-agencia-zaragoza", "Venta / vender sin inmobiliaria"),
            ("https://livendia.com/servicios/venta-piso-particular-sin-agencia", "Venta / vender sin inmobiliaria"),
            ("https://livendia.com/vender-piso-sin-inmobiliaria", "Venta / vender sin inmobiliaria"),
            (f"{PILLAR_VENDER_BASE}/barcelona", "Venta / vender sin inmobiliaria"),
            (f"{PILLAR_VENDER_BASE}/bilbao", "Venta / vender sin inmobiliaria"),
            (f"{PILLAR_VENDER_BASE}/granada", "Venta / vender sin inmobiliaria"),
            (f"{PILLAR_VENDER_BASE}/madrid", "Venta / vender sin inmobiliaria"),
            (f"{PILLAR_VENDER_BASE}/malaga", "Venta / vender sin inmobiliaria"),
            (f"{PILLAR_VENDER_BASE}/sevilla", "Venta / vender sin inmobiliaria"),
            (f"{PILLAR_VENDER_BASE}/valencia", "Venta / vender sin inmobiliaria"),
        ],
    ),
]

# Bloque único 50 URLs (orden exacto del usuario)
BLOQUE_UNICO: list[str] = [
    "https://livendia.com/gestoria/les-corts",
    "https://livendia.com/servicios/acompanamiento-alquiler-local",
    "https://livendia.com/servicios/acompanamiento-alquiler-local/hospitalet-de-llobregat",
    "https://livendia.com/servicios/acompanamiento-alquiler-local/madrid",
    "https://livendia.com/servicios/contrato-alquiler-local/barcelona-gracia",
    "https://livendia.com/servicios/contrato-alquiler-local/barcelona-les-corts",
    "https://livendia.com/servicios/contrato-alquiler-local/cornella-de-llobregat",
    "https://livendia.com/servicios/contrato-alquiler-local/hospitalet-de-llobregat",
    "https://livendia.com/servicios/contrato-alquiler-temporada-local/malaga",
    "https://livendia.com/servicios/contrato-alquiler-temporada-local/valencia",
    "https://livendia.com/servicios/contrato-arras-local/malaga",
    "https://livendia.com/servicios/contrato-arras-local/palma",
    "https://livendia.com/servicios/contrato-arras-local/sant-cugat-del-valles",
    "https://livendia.com/servicios/contrato-arras-local/sevilla",
    "https://livendia.com/servicios/contrato-arras-local/zaragoza",
    "https://livendia.com/servicios/contrato-entre-particulares-local",
    "https://livendia.com/servicios/contrato-entre-particulares-local/alicante",
    "https://livendia.com/servicios/contrato-entre-particulares-local/bilbao",
    "https://livendia.com/servicios/contrato-entre-particulares-local/cordoba",
    "https://livendia.com/servicios/contrato-entre-particulares-local/gijon",
    "https://livendia.com/servicios/contrato-entre-particulares-local/granada",
    "https://livendia.com/servicios/contrato-entre-particulares-local/valladolid",
    "https://livendia.com/servicios/contrato-entre-particulares-local/zaragoza",
    "https://livendia.com/servicios/reserva-de-compra",
    "https://livendia.com/servicios/revision-contrato-alquiler",
    "https://livendia.com/servicios/revision-documental-post-arras",
    "https://livendia.com/servicios/revision-documental-post-arras/cornella-de-llobregat",
    "https://livendia.com/servicios/revision-documental-post-arras/hospitalet-de-llobregat",
    "https://livendia.com/servicios/revision-documental-post-arras/les-corts",
    "https://livendia.com/servicios/revision-documental-post-arras/malaga",
    "https://livendia.com/servicios/revision-documental-post-arras/valencia",
    "https://livendia.com/servicios/vender-piso-sin-agencia",
    "https://livendia.com/servicios/vender-piso-sin-agencia-granada",
    "https://livendia.com/servicios/vender-piso-sin-agencia-zaragoza",
    "https://livendia.com/servicios/venta-piso-particular-sin-agencia",
    "https://livendia.com/vender-piso-sin-inmobiliaria",
    f"{PILLAR_VENDER_BASE}/barcelona",
    f"{PILLAR_VENDER_BASE}/bilbao",
    f"{PILLAR_VENDER_BASE}/granada",
    f"{PILLAR_VENDER_BASE}/madrid",
    f"{PILLAR_VENDER_BASE}/malaga",
    f"{PILLAR_VENDER_BASE}/sevilla",
    f"{PILLAR_VENDER_BASE}/valencia",
    "https://livendia.com/servicios/contrato-alquiler-habitacion/barcelona",
    "https://livendia.com/servicios/contrato-alquiler-habitacion/madrid",
    "https://livendia.com/servicios/contrato-alquiler-local/barcelona",
    "https://livendia.com/servicios/contrato-arras-local/asturias",
    "https://livendia.com/gestoria/sevilla",
    "https://livendia.com/ciudades/barcelona",
    "https://livendia.com/servicios/vender-piso-sin-agencia-barcelona",
]

HEADER_FILL = PatternFill("solid", fgColor="1A4FBF")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style_header_row(ws, cols: int = 2) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_category_sheet(wb, sheet_name: str, rows: list[Row]) -> None:
    ws = wb.create_sheet(sheet_name[:31])
    ws.append(["#", "URL", "Categoría"])
    style_header_row(ws, 3)
    for i, (url, cat) in enumerate(rows, start=1):
        ws.append([i, url, cat])
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 36
    ws.freeze_panes = "A2"


def main() -> None:
    OUT.parent.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()

    ws0 = wb.active
    ws0.title = "Resumen"
    ws0.append(["Livendia — Solicitud indexación GSC"])
    ws0.append(["Generado", str(date.today())])
    ws0.append(["Total URLs bloque único", len(BLOQUE_UNICO)])
    ws0.append([])
    ws0.append(["Hoja", "Descripción"])
    ws0.append(["Todas 50 URLs", "Lista maestra copy-paste GSC (orden exacto)"])
    for name, rows in SHEETS:
        ws0.append([name, f"{len(rows)} URLs por categoría"])
    ws0["A1"].font = Font(bold=True, size=14)
    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 48

    ws_master = wb.create_sheet("Todas 50 URLs")
    ws_master.append(["#", "URL"])
    style_header_row(ws_master, 2)
    for i, url in enumerate(BLOQUE_UNICO, start=1):
        ws_master.append([i, url])
    ws_master.column_dimensions["A"].width = 5
    ws_master.column_dimensions["B"].width = 72
    ws_master.freeze_panes = "A2"

    for sheet_name, rows in SHEETS:
        write_category_sheet(wb, sheet_name, rows)

    wb.save(OUT)
    print(OUT.resolve())
    print(f"URLs bloque único: {len(BLOQUE_UNICO)}")


if __name__ == "__main__":
    main()
